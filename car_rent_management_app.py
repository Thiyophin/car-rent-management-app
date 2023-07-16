import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from PIL import Image, ImageTk
import tkinter.messagebox as messagebox
import openpyxl
import shutil
import os


current_image_path = None
selected_serial_number = None
preview_label = None
name_entry = None
phone_entry = None
address_entry = None
car_entry = None
start_entry = None
end_entry = None
time_entry = None
issues_entry = None
advance_entry = None
pending_entry = None
table_view = None


def browse_image():
    global current_image_path
    file_path = filedialog.askopenfilename(
        filetypes=[("Image files", "*.png;*.jpg;*.jpeg")])
    if file_path:
        current_image_path = file_path
        hide_preview()
        load_image(file_path)
    else:
        hide_preview()


def save_customer():
    global name_entry, phone_entry, address_entry, car_entry, start_entry, end_entry, time_entry, issues_entry, advance_entry, pending_entry, table_view

    customer_name = name_entry.get()
    phone_number = phone_entry.get()
    address = address_entry.get()
    car_name = car_entry.get()
    start_date = start_entry.get()
    end_date = end_entry.get()
    time = time_entry.get()
    issues = issues_entry.get()
    advance_payment = advance_entry.get()
    pending_payment = pending_entry.get()

    # Check if any field is empty
    if not (customer_name and phone_number and address and car_name and start_date and end_date and time and issues and advance_payment and pending_payment):
        messagebox.showerror("Error", "Please fill in all the fields.")
        return

    save_dir = "Customer Details Folder"
    excel_file_path = os.path.join(save_dir, "customer_details.xlsx")
    images_dir = "Car Images Folder"

    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    if not os.path.exists(images_dir):
        os.makedirs(images_dir)

    if os.path.isfile(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Customer Details"
        sheet["A1"] = "Serial Number"
        sheet["B1"] = "Customer Name"
        sheet["C1"] = "Phone Number"
        sheet["D1"] = "Address"
        sheet["E1"] = "Car Name"
        sheet["F1"] = "Start Date"
        sheet["G1"] = "End Date"
        sheet["H1"] = "Time"
        sheet["I1"] = "Issues"
        sheet["J1"] = "Advance Payment"
        sheet["K1"] = "Pending Payment"
    # Get the existing serial number or start from 1
    serial_number = sheet.max_row if sheet.max_row > 1 else 1

    row = (serial_number, customer_name, phone_number, address, car_name,
           start_date, end_date, time, issues, advance_payment, pending_payment)
    sheet.append(row)

    workbook.save(excel_file_path)

    if current_image_path:
        image_file_name = f"{customer_name}_{serial_number}.png"
        shutil.copyfile(current_image_path, os.path.join(
            images_dir, image_file_name))
        messagebox.showinfo("Success", "Customer details saved successfully.")
    else:
        messagebox.showwarning("Warning", "No image selected.")

    name_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    address_entry.delete(0, tk.END)
    car_entry.delete(0, tk.END)
    start_entry.delete(0, tk.END)
    end_entry.delete(0, tk.END)
    time_entry.delete(0, tk.END)
    issues_entry.delete(0, tk.END)
    advance_entry.delete(0, tk.END)
    pending_entry.delete(0, tk.END)
    hide_preview()

    # Update table with saved customer details
    update_table()


def update_table():
    global table_view

    excel_file_path = os.path.join(
        "Customer Details Folder", "customer_details.xlsx")
    images_dir = "Car Images Folder"

    if not os.path.isfile(excel_file_path):
        return

    # Clear existing items in the table
    table_view.delete(*table_view.get_children())

    # Load customer details from Excel
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Configure the table columns
    table_view.column("#0", width=100, anchor="center")
    table_view.column("Serial Number", width=100, anchor="center")
    table_view.column("Customer Name", width=150, anchor="center")
    table_view.column("Phone Number", width=150, anchor="center")
    table_view.column("Address", width=150, anchor="center")
    table_view.column("Car Name", width=150, anchor="center")
    table_view.column("Start Date", width=150, anchor="center")
    table_view.column("End Date", width=150, anchor="center")
    table_view.column("Time", width=150, anchor="center")
    # Change column index to 9
    table_view.column("Issues", width=150, anchor="center")
    table_view.column("Advance Payment", width=150, anchor="center")
    table_view.column("Pending Payment", width=150, anchor="center")

    # Add details to the table
    for row in sheet.iter_rows(min_row=2, values_only=True):
        serial_number = row[0]
        customer_name = row[1]
        phone_number = row[2]
        address = row[3]
        car_name = row[4]
        start_date = row[5]
        end_date = row[6]
        time = row[7]
        issues = row[8]
        advance_payment = row[9]
        pending_payment = row[10]

        # Add details to the table
        table_view.insert("", tk.END, values=(
            serial_number, customer_name, phone_number, address, car_name, start_date, end_date, time, issues, advance_payment, pending_payment))


def load_image(file_path):
    global preview_label
    image = Image.open(file_path)
    image = image.resize((200, 200))
    image_preview = ImageTk.PhotoImage(image)
    preview_label.configure(image=image_preview)
    preview_label.image = image_preview
    show_preview()


def show_preview():
    preview_label.grid(row=11, column=0, columnspan=2, padx=10, pady=5)


def hide_preview():
    preview_label.grid_remove()


def clear_entries():
    global name_entry, phone_entry, address_entry, car_entry, start_entry, end_entry, time_entry, issues_entry, advance_entry, pending_entry

    name_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    address_entry.delete(0, tk.END)
    car_entry.delete(0, tk.END)
    start_entry.delete(0, tk.END)
    end_entry.delete(0, tk.END)
    time_entry.delete(0, tk.END)
    issues_entry.delete(0, tk.END)
    advance_entry.delete(0, tk.END)
    pending_entry.delete(0, tk.END)


def login():
    username = username_entry.get()
    password = password_entry.get()

    if username == "admin" and password == "password":
        messagebox.showinfo("Success", "Login successful!")
        login_window.destroy()
        open_main_window()
    else:
        messagebox.showerror("Error", "Invalid username or password.")


def open_main_window():
    global name_entry, phone_entry, address_entry, car_entry, start_entry, end_entry, time_entry, issues_entry, advance_entry, pending_entry, preview_label, table_view

    root = tk.Tk()
    root.title("Car Rental")
    icon_path = "rent-a-car_10754245.ico"  # Replace with the actual path to your .png icon
    root.iconbitmap(icon_path)

    root.configure(bg="#90EE90")

    details_frame = tk.Frame(root, bg="#f2f2f2")
    details_frame.pack(padx=10, pady=10)

    name_label = tk.Label(details_frame, text="Customer Name:",
                          font=("Arial", 12), bg="#f2f2f2")
    name_label.grid(row=0, column=0, sticky="w")
    name_entry = tk.Entry(details_frame, font=("Arial", 12))
    name_entry.grid(row=0, column=1, padx=10)

    phone_label = tk.Label(
        details_frame, text="Customer Phone Number:", font=("Arial", 12), bg="#f2f2f2")
    phone_label.grid(row=1, column=0, sticky="w")
    phone_entry = tk.Entry(details_frame, font=("Arial", 12))
    phone_entry.grid(row=1, column=1, padx=10)

    address_label = tk.Label(
        details_frame, text="Customer Address:", font=("Arial", 12), bg="#f2f2f2")
    address_label.grid(row=2, column=0, sticky="w")
    address_entry = tk.Entry(details_frame, font=("Arial", 12))
    address_entry.grid(row=2, column=1, padx=10)

    car_label = tk.Label(details_frame, text="Car Name:",
                         font=("Arial", 12), bg="#f2f2f2")
    car_label.grid(row=3, column=0, sticky="w")
    car_entry = tk.Entry(details_frame, font=("Arial", 12))
    car_entry.grid(row=3, column=1, padx=10)

    start_label = tk.Label(details_frame, text="Start Date:",
                           font=("Arial", 12), bg="#f2f2f2")
    start_label.grid(row=4, column=0, sticky="w")
    start_entry = tk.Entry(details_frame, font=("Arial", 12))
    start_entry.grid(row=4, column=1, padx=10)

    end_label = tk.Label(details_frame, text="End Date:",
                         font=("Arial", 12), bg="#f2f2f2")
    end_label.grid(row=5, column=0, sticky="w")
    end_entry = tk.Entry(details_frame, font=("Arial", 12))
    end_entry.grid(row=5, column=1, padx=10)

    time_label = tk.Label(details_frame, text="Time:",
                          font=("Arial", 12), bg="#f2f2f2")
    time_label.grid(row=6, column=0, sticky="w")
    time_entry = tk.Entry(details_frame, font=("Arial", 12))
    time_entry.grid(row=6, column=1, padx=10)

    issues_label = tk.Label(
        details_frame, text="Car Issues:", font=("Arial", 12), bg="#f2f2f2")
    issues_label.grid(row=7, column=0, sticky="w")
    issues_entry = tk.Entry(details_frame, font=("Arial", 12))
    issues_entry.grid(row=7, column=1, padx=10)

    advance_label = tk.Label(
        details_frame, text="Advance Payment:", font=("Arial", 12), bg="#f2f2f2")
    advance_label.grid(row=8, column=0, sticky="w")
    advance_entry = tk.Entry(details_frame, font=("Arial", 12))
    advance_entry.grid(row=8, column=1, padx=10)

    pending_label = tk.Label(
        details_frame, text="Pending Payment:", font=("Arial", 12), bg="#f2f2f2")
    pending_label.grid(row=9, column=0, sticky="w")
    pending_entry = tk.Entry(details_frame, font=("Arial", 12))
    pending_entry.grid(row=9, column=1, padx=10)

    image_label = tk.Label(details_frame, text="Customer ID:",
                           font=("Arial", 12), bg="#f2f2f2")
    image_label.grid(row=10, column=0, sticky="w")

    browse_button = tk.Button(details_frame, text="Browse", font=(
        "Arial", 12), command=browse_image)
    browse_button.grid(row=10, column=1, pady=5)

    preview_label = tk.Label(details_frame, width=200,
                             height=200, bg="white", relief="solid")
    preview_label.grid(row=11, column=0, columnspan=2, padx=10, pady=5)
    hide_preview()

    save_button = tk.Button(root, text="Save", font=(
        "Arial", 12), command=save_customer)
    save_button.pack(pady=10)

    style = ttk.Style()
    style.theme_use("clam")

    style.configure("Treeview", background="#f2f2f2",
                    foreground="black", rowheight=25, fieldbackground="#f2f2f2")
    style.map("Treeview", background=[("selected", "#347083")])

    scrollbar = ttk.Scrollbar(root)
    scrollbar.pack(side="right", fill="y")

    table_view = ttk.Treeview(root, columns=("Serial Number", "Customer Name", "Phone Number", "Address", "Car Name", "Start Date",
                              "End Date", "Time", "Issues", "Advance Payment", "Pending Payment"), show="headings", style="Treeview", yscrollcommand=scrollbar.set)
    table_view.heading("Serial Number", text="Serial No")
    table_view.heading("Customer Name", text="Customer Name")
    table_view.heading("Phone Number", text="Phone Number")
    table_view.heading("Address", text="Address")
    table_view.heading("Car Name", text="Car Name")
    table_view.heading("Start Date", text="Start Date")
    table_view.heading("End Date", text="End Date")
    table_view.heading("Time", text="Time")
    table_view.heading("Issues", text="Car Issues")
    table_view.heading("Advance Payment", text="Advance Payment")
    table_view.heading("Pending Payment", text="Pending Payment")

    scrollbar.config(command=table_view.yview)
    table_view.pack(expand=True, fill="both")

    update_table()

    root.mainloop()


login_window = tk.Tk()
login_window.title("Login")
login_window.configure(bg="#90EE90")

login_frame = tk.Frame(login_window, bg="#f2f2f2")
login_frame.pack(padx=10, pady=10)

username_label = tk.Label(login_frame, text="Username:",
                          font=("Arial", 12), bg="#f2f2f2")
username_label.grid(row=0, column=0, sticky="w")
username_entry = tk.Entry(login_frame, font=("Arial", 12))
username_entry.grid(row=0, column=1, padx=10)

password_label = tk.Label(login_frame, text="Password:",
                          font=("Arial", 12), bg="#f2f2f2")
password_label.grid(row=1, column=0, sticky="w")
password_entry = tk.Entry(login_frame, show="*", font=("Arial", 12))
password_entry.grid(row=1, column=1, padx=10)

login_button = tk.Button(login_frame, text="Login",
                         font=("Arial", 12), command=login)
login_button.grid(row=2, column=0, columnspan=2, pady=10)

login_window.mainloop()
